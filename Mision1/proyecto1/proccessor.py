# LA LOGICA DEL NEGOCIO. LO QUE HACE EL SISTEMA
# PROCESA LA INFORMACIÓN

import re
from openpyxl import load_workbook
# =======================================================
# FUNCION clean_id
# FUNCION ELIMINA CATACTERES NO NUMERICOS DE UN DOCUMENTO
# "cc.111.530.476" = "111530476"
# =======================================================
def clean_id(value):
    if  value is None:
        return ""
    return re.sub(r'\D', '', str(value))
# ======================================
# FUNCIÓN merge_name
# Une nombre y apellido en un solo campo
# ======================================
def merge_name(name, last_name):
    if name is None:
        name = ""
    if last_name is None:
        last_name = ""
    return f"{name} {last_name}".strip()
# ===========================================================================
# ABRE UN ARCHIVO DE EXCEL Y ENTRAR A UNA HOJA ESPECÍFICA QUE SE LLAMA “DATOS”.
# ============================================================================
def process_excel(path):
    #Acceso a la hoja llamada "Datos"
    wb= load_workbook(path)
    ws= wb["Datos"]
    #Recorrer todas las filas desde la fila 2
    for row in range(2, ws.max_row + 1):
        #Columna D = Limpia el documento de la columna A
        ws[f"D{row}"]= clean_id(ws[f"A{row}"].value)
        #Columna E = Une nombre y apellido en un solo campo
        ws[f"E{row}"]= merge_name(ws[f"B{row}"].value, ws[f"C{row}"].value)